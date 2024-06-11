###############################
# CodeX Learn python          #
# Assignment 7                #
# Writing to Excel            #
# 20/5/2024                   #
###############################
import openpyxl as opx
import matplotlib.pyplot as plt
import pandas as pd

# Debug Mode - set to True to print debug messages
debug_mode = True

attendance_record = []


#############################
# Create a new excel file
# and add some data
#
#############################


#############################
# Create a new excel file
# and add some data
#
#############################
def create_excel_file(data, filename):
  workbook = opx.Workbook()
  sheet = workbook.active
  sheet.title = "Data"
  
  # Add headers to the sheet
  headers = ["Name", "Date", "Status" ]
  sheet.append(headers)
    
  for row in data:
    sheet.append(row)
    
  workbook.save(filename)
  
  print(f"The excel file {filename} was created")
  
#############################
# Create an attendance record
#
# @name - name of attendee
# @date - date of attendance
# @status - status (present/absent)
#
#############################
def add_attendance_record(name, date, status):
 attendance_record.update({"Name": name, "Date": date, "Status": status})
 
 if debug_mode == True:
  print(f"Name: {name}, Date: {date}, Status: {status}")
   
#############################
# open an excel file
#
#############################
def read_excel_file(filename):
  workbook = opx.load_workbook(filename)
  sheet = workbook.active
  
  data = []
  for row in sheet.iter_rows(values_only=True):
    data.append(row)
  
  return data
  
  
  
#############################
# Plot a graph using our data
#
#
#############################
def visualise_data(data):
  # Exclude the header so start at 1
  name = [row[0] for row in data[1:] ]
  date = [row[1] for row in data[1:] ]
  status = [row[2] for row in data[1:] ]
  
  plt.figure(figsize=(10,6))


  # Plot a line graph with a o at each data point
  plt.plot(date, status, marker='x', linestyle='-', color='#0000FF')
  plt.title('Daily Attendance')
  plt.xlabel('Date')
  plt.ylabel('Status')
  plt.grid(True)
  plt.show()
  
#############################
# Main()
#
#
#############################
def main():
  
  # Excel Filename 
  filename = "attendance_record.xlsx"

        
  # Print he request to add data
  print("Welcome to the attendance register")


  # loop until break
  while True:
  
    # Print he request to add data
    request=input("Please enter \n'a' to add an attendance record, \n'v' to view an attendance record,\n'q' to quit !\n")
  
    if request == 'a':
      # Print input details
      name = input("Please enter attendee name:")
      date = input(f"Please enter the date (DD/MM/YY):")
      status = input(f"Please enter a status (Attended/Absent):")

      # Create a new row of data and append to record
      row = [name, date, status]
      attendance_record.append(row)

      if debug_mode == True:
        print(f"Name: {name}, Date: {date}, Status: {status}")

      
    elif request == 'v':
      
      # First write all data to file and then display
      create_excel_file(attendance_record, filename)
      dt = read_excel_file(filename)
  
      print("Data from excel file: ")
      for row in dt:
        print(row)
      
      visualise_data(dt)
      
    elif request == 'q':
      print("Quitting")
      break
    else:
      print("Invalid input, try again")




    
if __name__ == "__main__":
  main()
  
  
  